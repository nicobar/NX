from openpyxl import load_workbook
import ciscoconfparse as c
import re
# import ipaddr
import itertools


#############################################
################# VARIABLES #################
#############################################

SWITCH = 'NAOSW133'
SHEET = SWITCH
BASE_DIR = '/Users/aspera/Documents/Clienti/VF-2017/NMP/NA1C/' + SWITCH + '/Stage_3/'


INPUT_XLS = BASE_DIR + SWITCH + '_OUT_DB_OPT.xlsx'
OSW_CFG_TXT = BASE_DIR + SWITCH + '.txt'
OSWVCE_CFG_TXT = BASE_DIR + SWITCH + 'VCE' +'.txt'
# OSWVCE_CFG_IF_TXT = BASE_DIR + SWITCH + 'VCE' + '-' + 'IF' + '.txt'
# OSWVCE_CFG_IF_VLAN_TXT = BASE_DIR + SWITCH + 'VCE' + '-' + 'IF' + '-' + 'VLAN' + '.txt'
# OSWVCE_CFG_IF_VLAN_SVI_TXT = BASE_DIR + SWITCH + 'VCE' + '-' + 'IF' + '-' + 'VLAN' + '-' + 'SVI' + '.txt'

#############################################
################ FUNCTIONS ##################
#############################################

def atoi(text):
    ''' from string to int'''
    return int(text) if text.isdigit() else text

def natural_keys(text):
    '''
    alist.sort(key=natural_keys) sorts in human order
    http://nedbatchelder.com/blog/200712/human_sorting.html
    (See Toothy's implementation in the comments)
    '''
    return [ atoi(c) for c in re.split('(\d+)', text) ]

def get_col_N3048(ws,col):
    ''' Take a worksheet, return column "col" as lists conditioned to col = 6 == "N3048" '''
    NEXUS_AP_COL = 6
    return [str(ws.cell(row = r, column = col).value) for r in range(2,ws.max_row+1) if ws.cell(row = r, column = NEXUS_AP_COL).value == 'N3048' ]

def get_col_N9508(ws,col):
    ''' Take a worksheet, return column "col" as lists conditioned to col = 6 == "N3048" '''
    NEXUS_AP_COL = 6
    return [str(ws.cell(row = r, column = col).value) for r in range(2,ws.max_row+1) if ws.cell(row = r, column = NEXUS_AP_COL).value != 'N3048' ]

def get_if_from_xls():
    ''' Return column col as list '''
    wb_r = load_workbook(INPUT_XLS)
    ws_r = wb_r.get_sheet_by_name(SHEET)
    SRC_IF_COL = 1
    if_N3048 = get_col_N3048(ws_r,SRC_IF_COL)
    if_N9508 = get_col_N9508(ws_r,SRC_IF_COL)
    if_N3048.sort(key=natural_keys)
    if_N9508.sort(key=natural_keys)
    return (if_N9508,if_N3048)

def get_if_from_cfg():
    parse = c.CiscoConfParse(OSW_CFG_TXT)
    intf_obj_list = parse.find_objects(r'^interface .*Ethernet')
    
    a = [obj.text for obj in intf_obj_list]
    a.sort(key=natural_keys)
    return a
    
def get_vlan_from_cfg():
    parse = c.CiscoConfParse(OSW_CFG_TXT)
    vlan_obj_list = parse.find_objects(r'^vlan \d+')
    
    return [obj.text.split(' ')[1] for obj in vlan_obj_list]

def get_vlan_from_xls():
    
    a_N9508 = set()
    a_N3048 = set()
    wb_r = load_workbook(INPUT_XLS)
    ws_r = wb_r.get_sheet_by_name(SHEET)
    VLAN_COL = 4

    lst_N9508 = get_col_N9508(ws_r,VLAN_COL)
    lst_N3048 = get_col_N3048(ws_r,VLAN_COL)
    
    for elem_N9508 in lst_N9508:
        if ',' in elem_N9508:
            b_N9508 = elem_N9508.split(',')
            for elem2 in b_N9508:
                a_N9508.add(elem2)
        else:
            a_N9508.add(elem_N9508)
            
    for elem_N3048 in lst_N3048:            
        if ',' in elem_N3048:
            b_N3048 = elem_N3048.split(',')
            for elem2 in b_N3048:
                a_N3048.add(elem2)
        else:
            a_N3048.add(elem_N3048)

    lst2_N9508 = list(a_N9508)
    lst2_N3048 = list(a_N3048)
    lst2_N9508.sort(key=natural_keys)
    lst2_N3048.sort(key=natural_keys)
    return (lst2_N9508,lst2_N3048)

def get_svi_from_cfg():
    parse = c.CiscoConfParse(OSW_CFG_TXT)
    svi_obj_list = parse.find_objects(r'^interface Vlan')
    
#    lst = [obj.text for obj in svi_obj_list]       # this get ["interface VlanX",...]
#  lst2 = [elem.split(' ')[1] for elem in lst]     # this get ["VlanX",...]
#  lst3 = [re.findall('\d+',x)[0] for x in lst2]   # this get ["X",...]
#  lst3.sort(key=natural_keys)   
#    return lst3

    lst = [re.findall(r'^interface Vlan(\d+)',svi_obj.text)[0] for svi_obj in svi_obj_list] 
    return lst

def get_svi_on_device(vlanxls, svi_from_cfg):
    a = [x for x in svi_from_cfg if x in vlanxls]
    a.sort(key=natural_keys)
    return a

def get_list_not_to_be_migrated(ifxls,ifcfg):
    a = set(ifxls)
    b = set(ifcfg)
    c = b-a
    d=list(c)
    if len(d) > 0:
        d.sort(key=natural_keys)
        return d 
    else:
        return []
    
def write_normalized_OSWVCE_cfg(if_ntbm_N9508, vlan_ntbm_N9508, svi_ntbm_N9508):
    parse = c.CiscoConfParse(OSW_CFG_TXT)

    # FIRST BLOCK --> INTERFACES
    intf_obj_list = parse.find_objects(r'^interface .*Ethernet')
    
    for intf_obj in intf_obj_list:
        if intf_obj.text in if_ntbm_N9508:
            intf_obj.delete()
            
    parse.commit()
    #parse.save_as(OSWVCE_CFG_IF_TXT)
    intf_obj_list = parse.find_objects(r'^interface .*Ethernet')
    cfg_intf_list = [intf_obj.ioscfg + ['!'] for intf_obj in intf_obj_list]
    cfg_intf =  list(itertools.chain.from_iterable(cfg_intf_list))
    print "done if_cfg"
    
    
    # SECOND BLOCK --> VLANS
    vlan_obj_list = parse.find_objects(r'^vlan \d+$')
    
    for vlan_obj in vlan_obj_list:
        vlan = vlan_obj.text
        if re.findall(r'^vlan (\d+)$',vlan)[0] in vlan_ntbm_N9508:
            vlan_obj.delete()
            
    parse.commit()
    vlan_obj_list = parse.find_objects(r'^vlan \d+$')
    cfg_vlan_list = [vlan_obj.ioscfg + ['!'] for vlan_obj in vlan_obj_list]
    cfg_vlan =  list(itertools.chain.from_iterable(cfg_vlan_list))
    print "done vlan_cfg"
    
    # THIRD BLOCK --> SVI
    svi_obj_list = parse.find_objects(r'^interface Vlan')
    
    for svi_obj in svi_obj_list:
        svi = svi_obj.text
        num_svi = re.findall(r'^interface Vlan(\d+)$',svi)[0]
        if num_svi in svi_ntbm_N9508:
            svi_obj.delete()
            
    parse.commit()    
    svi_obj_list = parse.find_objects(r'^interface Vlan')
    cfg_svi_list = [svi_obj.ioscfg + ['!'] for svi_obj in svi_obj_list]
    cfg_svi =  list(itertools.chain.from_iterable(cfg_svi_list))
    print "done vlan_cfg"
    
    cfg = cfg_vlan + cfg_intf + cfg_svi 
    parse_out =  c.CiscoConfParse(cfg)
    parse_out.save_as(OSWVCE_CFG_TXT)
    print "done write"

# def write_normalized_if_OSWVCE_cfg(if_ntbm_N9508):
#     parse = c.CiscoConfParse(OSW_CFG_TXT)
# 
#     
#     intf_obj_list = parse.find_objects(r'^interface .*Ethernet')
#     
#     for intf_obj in intf_obj_list:
#         if intf_obj.text in if_ntbm_N9508:
#             intf_obj.delete()
#             
#     parse.commit()
#     parse.save_as(OSWVCE_CFG_IF_TXT)
#     print "fine write_normalized_OSWVCE_if_cfg"
#     
#     
#     
# def write_normalized_vlan_OSWVCE_cfg(vlan_ntbm_N9508):
#     parse = c.CiscoConfParse(OSWVCE_CFG_IF_TXT)
# 
#     
#     intf_obj_list = parse.find_objects(r'^vlan \d+$')
#     
#     for intf_obj in intf_obj_list:
#         vlan = intf_obj.text
#         if re.findall(r'^vlan (\d+)$',vlan)[0] in vlan_ntbm_N9508:
#             intf_obj.delete()
#             
#     parse.commit()        
#     parse.save_as(OSWVCE_CFG_IF_VLAN_TXT)
#     print "fine write_normalized_OSWVCE_vlan_cfg"
# 
#     
# def write_normalized_svi_OSWVCE_cfg(svi_ntbm_N9508):
#     parse = c.CiscoConfParse(OSWVCE_CFG_IF_VLAN_TXT)
# 
#     
#     svi_obj_list = parse.find_objects(r'^interface Vlan')
#     
#     for svi_obj in svi_obj_list:
#         svi = svi_obj.text
#         num_svi = re.findall(r'^interface Vlan(d+)$',svi)[0]
#         if num_svi in svi_ntbm_N9508:
#             svi_obj.delete()
#             
#     parse.commit()        
#     parse.save_as(OSWVCE_CFG_IF_VLAN_SVI_TXT)
#     print "fine write_normalized_OSWVCE_svi_cfg"

#############################################
################### MAIN ####################
#############################################


if_xls_N9508, if_xls_N3048 = get_if_from_xls()
if_cfg = get_if_from_cfg()

print "if_xls_N9508 = ", if_xls_N9508
print "if_xls_N3048 = ", if_xls_N3048
print "if_cfg = ", if_cfg

if_not_to_be_migrated_N9508 = get_list_not_to_be_migrated(if_xls_N9508, if_cfg)
if_not_to_be_migrated_N3048 = get_list_not_to_be_migrated(if_xls_N3048, if_cfg)
print "if_not_to_be_migrated_N9508 = " , if_not_to_be_migrated_N9508
print "if_not_to_be_migrated_N3048 = " , if_not_to_be_migrated_N3048

vlan_xls_N9508, vlan_xls_N3048 = get_vlan_from_xls()
vlan_cfg = get_vlan_from_cfg()

print "vlan_xls_N9508 = ", vlan_xls_N9508
print "vlan_xls_N3048 = ", vlan_xls_N3048


vlan_not_to_be_migrated_N9508 = get_list_not_to_be_migrated(vlan_xls_N9508, vlan_cfg)
vlan_not_to_be_migrated_N3048 = get_list_not_to_be_migrated(vlan_xls_N3048, vlan_cfg)
print "vlan_not_to_be_migrated_N9508 = " , vlan_not_to_be_migrated_N9508
print "vlan_not_to_be_migrated_N3048 = " , vlan_not_to_be_migrated_N3048


svi_from_cfg = get_svi_from_cfg()
svi_on_N9508 = get_svi_on_device(vlan_xls_N9508, svi_from_cfg)
svi_on_N3048 = get_svi_on_device(vlan_xls_N3048, svi_from_cfg)

print "svi_from_cfg = " , svi_from_cfg
print "svi_on_N9508 = " , svi_on_N9508
print "svi_on_N3048 = " , svi_on_N3048

svi_not_to_be_migrated_N9508 = get_list_not_to_be_migrated(svi_on_N9508, svi_from_cfg)
svi_not_to_be_migrated_N3048 = get_list_not_to_be_migrated(svi_on_N3048, svi_from_cfg)

print "svi_not_to_be_migrated_N9508 = ", svi_not_to_be_migrated_N9508
print "svi_not_to_be_migrated_N3048 = ", svi_not_to_be_migrated_N3048

write_normalized_OSWVCE_cfg(if_not_to_be_migrated_N9508, vlan_not_to_be_migrated_N9508, svi_not_to_be_migrated_N9508)
# write_normalized_if_OSWVCE_cfg(if_not_to_be_migrated_N9508)
# write_normalized_vlan_OSWVCE_cfg(vlan_not_to_be_migrated_N9508)
# write_normalized_svi_OSWVCE_cfg(svi_not_to_be_migrated_N9508)
#write_normalized_OSWVSW_cfg(if_not_to_be_migrated_N3048, vlan_not_to_be_migrated_N3048, svi_not_to_be_migrated_N3048)