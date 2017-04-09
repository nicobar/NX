from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill
import ciscoconfparse as c
import re


#############################################
################# VARIABLES #################
#############################################

SWITCH = 'NAOSW133'
SHEET = SWITCH
BASE_DIR = '/Users/aspera/Documents/Clienti/VF-2017/NMP/NA1C/' + SWITCH + '/Stage_1/'


INPUT_XLS = BASE_DIR + SWITCH + '_DB_MIGRATION.xlsx'
OUTPUT_XLS = BASE_DIR + SWITCH + '_OUT_DB.xlsx'
OSW_CFG_TXT = BASE_DIR + SWITCH + '.txt'


# +-----0-A------+-----1-B------+------2-C------+---3-D--+---4-E-+-----5-F----+-------6-G---+-------7-H---------+-------8-I-----+-------9-J-----+-------10-K-----+----11-L----+-----12-M-------+---13-N----------+---14-N----------+---15-N----------+   
# +--SRC_OSW_IF--+--DST_VCE_IF--+--Access-Type--+--VLAN--+--QoS--+--Nexus_AP--+--Member/PO--+-----Descr---------+----Duplex-----+-----Speed-----+---Media Type---+---Action---+---Root-Guard---+---System-type---+---Check_Descr---+----Temp---------+ 
#                                   |                                    
#                                   |
#                                   +-- Access, trunk, infra

def get_string_from_range_to_list(range_str):
    ''' Takes '1-4' and Returns "1,2,3,4" '''
        
    help_list = range_str.split('-')
    start = int(help_list[0])
    stop = int(help_list[1])
    l = range(start,stop+1)
    stringed_list = [str(x) for x in l]    
    s = ','.join(stringed_list)
    return s

def get_allowed_vlan_list(if_cfg, SEL):
    ''' Get interface configuration block as a list 
        and returns a list (SEL = 'LIST') or string (SEL = "STRING")  
        of trunk allowed VLANS ''' 
    
    s = ''
    
    for line in if_cfg:
        
        if line[:30] == " switchport trunk allowed vlan":
            if line[:30] == " switchport trunk allowed vlan" and line[31:34] !="add":
                help_string = line[31:]
                help_string = str.rstrip(help_string)
                
                help_list = help_string.split(',')
                
                for elem in help_list:
                    if re.findall('-', elem):
                        s = s + get_string_from_range_to_list(elem) + ','
                    else:
                        s = s + elem + ','
                
            elif line[:34] == " switchport trunk allowed vlan add":
                help_string = line[35:]
                help_string = str.rstrip(help_string)
                
                help_list = help_string.split(',')
                
                for elem in help_list:
                    if re.findall('-', elem):
                        s = s + get_string_from_range_to_list(elem) + ','
                    else:
                        s = s + elem + ','
        else:
            continue
    s1 = s[:-1]
    if SEL == "STRING":
        return s1
    elif SEL == "LIST":
        return s1.split(',') 

def get_access_vlan(if_cfg):
    ''' Get interface configuration block as a list 
        and returns the access VLAN as integer ''' 
    
    for line in if_cfg:    
        if line[:23] == " switchport access vlan":
            access_vlan = int(line[24:])
            break
    return access_vlan     
   
def description_are_equals(desc_from_xls, if_cfg):
    ''' Get interface configuration block as a list + string desc_from_xls
        and returns True is are equal, False otherwise ''' 
 
    
    for elem in if_cfg:
        if elem[1:12] == 'description':
            desc_from_cfg = str.strip(elem[13:])
            if desc_from_xls == desc_from_cfg:
                return True
            else:
                return False
        
def get_channel_group(if_cfg):
    ''' Get interface configuration block as a list 
        and returns the channel-group id as integer ''' 
    
    for elem in if_cfg:
        if elem[1:14] == "channel-group":
            ch_gr = int(re.findall(r'\d+',elem)[0])
    return ch_gr    

def colour_output_xlsx():
    
    wb = load_workbook(OUTPUT_XLS)
    ws = wb.get_sheet_by_name(SHEET)
     
    
    MAX_COL = ws.max_column-1
    MAX_COLUMN_COLOR = MAX_COL-2
    print 'MAX_COL = ', ws.max_column 

    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')      # To be Deleted
    orangeFill = PatternFill(start_color='FF8000', end_color='FF8000', fill_type='solid')   # To be Checked
    yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')   # To be Merged
   
   
    for row in ws.rows:
        src_if = row[0].value
               
          
        if src_if[:12] == "Port-channel" or str.isdigit(str(row[6].value)):           # if portchannel or member then orange
            for cell in row[0:MAX_COLUMN_COLOR]:
                cell.fill = orangeFill
        if row[MAX_COL].value == 'temp':
            row[MAX_COL].value = ''
        if row[MAX_COL].value == 'r':
            row[MAX_COL].value = ''
            for cell in row[0:MAX_COLUMN_COLOR]:
                cell.fill = redFill
        elif row[MAX_COL].value == 'n':
            row[MAX_COL].value = ''   
        elif row[13].value == "Decommissioned" or row[12].value == "Decomissioned" or row[12].value == "Spare" or row[13].value == "Monitoring":
            for cell in row[0:MAX_COLUMN_COLOR]:
                cell.fill = redFill
        elif row[13].value == "TBV" or row[13].value == "TBV-NC" or row[13].value == "Infra": # IF Can be Core-Router, Core-Switch or Routed
            for cell in row[0:MAX_COLUMN_COLOR]:
                cell.fill = orangeFill
        elif row[13].value == "Core-Router" or row[13].value == "Core-Switch":
            for cell in row[0:MAX_COLUMN_COLOR]:
                cell.fill = yellowFill
                
    wb.save(filename = OUTPUT_XLS)
    print "End F2"

def readin_xls_writeout_xls():
    
    header_out = ['SRC OSW IF', 'DST VCE IF', 'Access Type', 'VLAN', 'QoS', 'Nexus AP', 'Member/PO', 'Descr', 'Duplex', 'Speed', 'Media Type', 'Action', 'Root-Guard', 'System Type', 'Check Descr', 'temp']
    
    parse = c.CiscoConfParse(OSW_CFG_TXT)
    
    intf_obj_list = parse.find_objects(r'^interface')
    wb_r = load_workbook(INPUT_XLS)
    wb_w = Workbook()
    
    
    ws_r = wb_r.get_sheet_by_name(SHEET)
    ws_w = wb_w.create_sheet(index = 0, title = SHEET)
    
    MAX_COL = 15
    MAX_ROW = ws_r.max_row
    

    pinkFill = PatternFill(start_color='eeaaee', end_color='eeaaee', fill_type='solid')     # To be Verified
    greenFill =  PatternFill(start_color='a7bd2f', end_color='a7bd2f', fill_type='solid') 
    
    
    ws_w.append(header_out)
    
 
    ws_w.cell(row=MAX_ROW+1,column=MAX_COL+1, value = 10).value
    
    
    for row_r,row_w in zip(ws_r.rows,ws_w.rows):
        if row_r[0].value == "Device":
            continue
        intf_from_xls = 'interface ' + str.strip(str(row_r[3].value))
        #print row_r[3].value
        row_w[0].value = str(row_r[3].value)                                             # Copy interface (or row_r[4].value)
        row_w[5].value = str(row_r[13].value)                                            # Copy New-Nexus AP into Nexus AP
        #row_w[7].value = row_r[5].value                                                 # Copy Port Description
        row_w[8].value = str(row_r[8].value)                                             # Copy Duplex
        row_w[9].value = str(row_r[9].value)                                             # Copy Speed
        row_w[10].value = str(row_r[10].value)                                           # Copy Port Media_type
        row_w[11].value = str(row_r[17].value)                                           # Copy Action
        row_w[12].value = str(row_r[11].value)                                           # Copy Root_Guard    
        row_w[13].value = str(row_r[12].value) 
        row_w[15].value = 'n'

               
        for intf_obj in intf_obj_list:
            intf_cfg = intf_obj.ioscfg                                                   # IOS IF's Configuration 
            intf_from_cfg = str.strip(intf_cfg[0])                                       # First line is IF itself
            
            if intf_from_xls == intf_from_cfg:
                                                 
                if intf_from_xls[0:22] == "interface Port-channel":
                    po = int(re.findall(r'\d+',intf_from_xls)[0])
                    row_w[6].value = po
                    
                if intf_obj.has_child_with("switchport trunk allowed vlan"):
                        vlan_list_string = get_allowed_vlan_list(intf_cfg, "STRING")             
                        row_w[2].value = 'Trunk'             
                        row_w[3].value = vlan_list_string
                        
                        if intf_obj.has_child_with("channel-group"):
                            channel_group = get_channel_group(intf_cfg)
                            row_w[6].value = channel_group
                                
                elif intf_obj.has_child_with("switchport mode access"):   
                    row_w[2].value = 'Access'   
                    if intf_obj.has_child_with("switchport access vlan"):
                        access_vlan = get_access_vlan(intf_cfg)                  
#                        row_w[2].value = 'Access'             
                        row_w[3].value = str(access_vlan)
                        if access_vlan == 1 and row_r[12].value != 'Monitoring':
                            row_w[15].value = 'r'                                       # colour_output_xlsx() will fill red
                    else:
                        row_w[3].value = 1
                        if row_r[12].value != 'Monitoring':
                            row_w[15].value = 'r'                                       # colour_output_xlsx() will fill red

                                
                if intf_obj.has_child_with("description"):
                    if description_are_equals(str.strip(str(row_r[5].value)), intf_cfg):
                        row_w[7].value = str(row_r[5].value)
                        row_w[14].value = "Description unchanged"
                        row_w[14].fill = greenFill
                    else:
                        row_w[7].value = "INTERFACE TO BE CHECKED"
                        row_w[14].value = "Description CHANGED!!!"
                        row_w[14].fill = pinkFill
                        
            else:
                continue            
             
    ws_w.cell(row=MAX_ROW+1,column=MAX_COL+1, value = '')
    wb_w.save(filename = OUTPUT_XLS)
    print "End F1"

    
readin_xls_writeout_xls()
colour_output_xlsx()